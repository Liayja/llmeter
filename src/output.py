"""
终端打印 + 结果保存
"""
import json
import csv
from pathlib import Path
from datetime import datetime

try:
    from rich.console import Console
    from rich.table import Table
    RICH = True
except ImportError:
    RICH = False 

def print_report(m: dict, model: str, engine_type: str):
    """打印压测报告到终端"""
    title = f"📊 压测结果 — {engine_type}/{model}"

    if RICH:
        _print_rich(m, title)
    else:
        _print_plain(m, title)

def _print_rich(m: dict, title: str):
    console = Console()
    table = Table(title=title)
    table.add_column("指标", style="bold cyan")
    table.add_column("值", justify="right")
    table.add_row("总请求", str(m["total"]))
    table.add_row("✅ 成功", f"[green]{m['success']}[/green]")
    table.add_row("❌ 失败", f"[red]{m['fail']}[/red]")
    table.add_row("成功率", f"{m['success_rate']:.1f}%")
    table.add_row("QPS", f"{m['qps']:.2f} req/s")
    if m.get("total_output_tokens", 0) > 0:
        table.add_row("Token 吞吐", f"{m['tps']:.1f} tok/s")
    table.add_section()
    table.add_row("[bold]延迟 (Latency)[/bold]", "")
    for label, key in [("Avg", "latency_avg"), ("P50", "latency_p50"),
                        ("P95", "latency_p95"), ("P99", "latency_p99"),
                        ("Min", "latency_min"), ("Max", "latency_max")]:
        table.add_row(f"  {label}", f"{m[key]:.3f}s")
    if m.get("ttft_avg") is not None:
        table.add_section()
        table.add_row("[bold]首 Token 延迟 (TTFT)[/bold]", "")
        for label, key in [("Avg", "ttft_avg"), ("P50", "ttft_p50"), ("P95", "ttft_p95")]:
            table.add_row(f"  {label}", f"{m[key]:.3f}s")
    if m.get("tpot_avg") is not None:
        table.add_section()
        table.add_row("[bold]每 Token 耗时 (TPOT)[/bold]", "")
        for label, key in [("Avg", "tpot_avg"), ("P50", "tpot_p50"), ("P95", "tpot_p95")]:
            table.add_row(f"  {label}", f"{m[key]*1000:.1f} ms")
    if m.get("itl_avg") is not None:
        table.add_section()
        table.add_row("[bold]Token 间隔 (ITL)[/bold]", "")
        for label, key in [("Avg", "itl_avg"), ("P50", "itl_p50"), ("P95", "itl_p95"), ("Max", "itl_max")]:
            table.add_row(f"  {label}", f"{m[key]*1000:.1f} ms")
    if m.get("concurrency_peak") is not None:
        table.add_section()
        table.add_row("[bold]并发[/bold]", f"{m['concurrency_peak']} (目标 {m.get('concurrency_target', '?')})")
    if m["errors"]:
        table.add_section()
        table.add_row("[bold red]错误摘要[/bold red]", "")
        for i, e in enumerate(m["errors"][:5]):
            table.add_row(f"  #{i+1}", f"[red]{e['msg'][:100]}[/red]")
    console.print(table)

def _print_plain(m: dict, title: str):
    """纯文本输出"""
    print(f"\n{'='*60}")
    print(f" {title}")
    print(f"{'='*60}")
    print(f" 总请求: {m['total']}  |  ✅ 成功: {m['success']}  |  ❌ 失败: {m['fail']}")
    print(f" 成功率: {m['success_rate']:.1f}%  |  QPS: {m['qps']:.2f} req/s")
    if m.get("total_output_tokens", 0) > 0:
        print(f" Token 吞吐: {m['tps']:.1f} tok/s")
    print(f" 延迟 — Avg: {m['latency_avg']:.3f}s  P50: {m['latency_p50']:.3f}s  P95: {m['latency_p95']:.3f}s  P99: {m['latency_p99']:.3f}s")
    if m.get("ttft_avg") is not None:
        print(f" TTFT — Avg: {m['ttft_avg']:.3f}s  P50: {m['ttft_p50']:.3f}s  P95: {m['ttft_p95']:.3f}s")
    if m.get("tpot_avg") is not None:
        print(f" TPOT — Avg: {m['tpot_avg']*1000:.1f}ms  P50: {m['tpot_p50']*1000:.1f}ms  P95: {m['tpot_p95']*1000:.1f}ms")
    if m["errors"]:
        print(f"\n 错误摘要:")
        for i, e in enumerate(m["errors"][:5]):
            print(f"   #{i+1}: {e['msg'][:120]}")


def save_results(metrics: dict, raw_results: list, cfg: dict, engine_type: str):
    """保存结果到 JSON 文件"""
    now = datetime.now()
    date_dir = Path(cfg["output_dir"]) / now.strftime("%Y-%m-%d")
    date_dir.mkdir(parents=True, exist_ok=True)
    safe_model = cfg["model"].replace("/", "_").replace(":", "_")
    filename = f"{safe_model}_c{cfg['concurrency']}_n{cfg['total_requests']}_{now.strftime('%H%M%S')}.json"
    filepath = date_dir / filename
    output = {
        "meta": {
            "engine": engine_type,
            "model": cfg["model"],
            "base_url": cfg["base_url"],
            "concurrency": cfg["concurrency"],
            "total_requests": cfg["total_requests"],
            "prompt": cfg["prompt"][:200],
            "timestamp": now.isoformat(),
        },
        "summary": metrics,
        "details": raw_results,
    }
    filepath.write_text(json.dumps(output, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\n💾 结果已保存: {filepath}")


# ═══════════════════════════════════════════════════
# 多场景对比报告
# ═══════════════════════════════════════════════════

# 对比表列定义：(表头, 取值函数, 格式化函数)
_COMPARE_COLS = [
    ("场景",      lambda r: r["name"],                       str),
    ("并发",      lambda r: r["concurrency"],                str),
    ("请求数",    lambda r: r["total"],                      str),
    ("入Token",   lambda r: r.get("avg_input_tokens"),       lambda v: str(v) if v is not None else "—"),
    ("出Token",   lambda r: r.get("avg_output_tokens"),      lambda v: str(v) if v is not None else "—"),
    ("总Token",   lambda r: r.get("total_tokens"),             lambda v: f"{v:,}" if v is not None else "—"),
    ("成功率",    lambda r: r["success_rate"],               lambda v: f"{v:.1f}%"),
    ("QPS",       lambda r: r["qps"],                        lambda v: f"{v:.2f}"),
    ("TPS",       lambda r: r["tps"],                        lambda v: f"{v:.1f}"),
    ("延迟Avg",   lambda r: r["latency_avg"],                lambda v: f"{v:.3f}s"),
    ("延迟Std",   lambda r: r.get("latency_std"),             lambda v: f"{v:.3f}s" if v else "—"),
    ("延迟P50",   lambda r: r["latency_p50"],                lambda v: f"{v:.3f}s"),
    ("延迟P95",   lambda r: r["latency_p95"],                lambda v: f"{v:.3f}s"),
    ("延迟P99",   lambda r: r["latency_p99"],                lambda v: f"{v:.3f}s"),
    ("TTFT Avg",  lambda r: r.get("ttft_avg"),               lambda v: f"{v:.3f}s" if v is not None else "—"),
    ("TTFT P95",  lambda r: r.get("ttft_p95"),               lambda v: f"{v:.3f}s" if v is not None else "—"),
    ("TPOT Avg",  lambda r: r.get("tpot_avg"),               lambda v: f"{v*1000:.1f}ms" if v is not None else "—"),
    ("TPOT P95",  lambda r: r.get("tpot_p95"),               lambda v: f"{v*1000:.1f}ms" if v is not None else "—"),
    ("长尾比",    lambda r: (r["latency_p95"] / r["latency_p50"]) if r["latency_p50"] else 0,
                  lambda v: f"{v:.2f}" if v != 0 else "—"),
]


def _row_from_result(name: str, metrics: dict, concurrency: int) -> dict:
    """把单场景的 metrics 摊平成对比表的一行"""
    return {"name": name, "concurrency": concurrency, **metrics}


# 跨模型对比时最左侧的「模型」列
_MODEL_COL = ("模型", lambda r: r.get("model_label") or r.get("_model") or "-", str)


def _cols(show_model: bool):
    """返回对比表列定义；show_model=True 时最左加一列模型。"""
    return ([_MODEL_COL] + _COMPARE_COLS) if show_model else _COMPARE_COLS


def print_comparison(rows: list, model: str, *, show_model: bool = False):
    """打印多场景对比表到终端"""
    title = f"📊 多场景压测对比 — {model}"
    cols = _cols(show_model)

    if RICH:
        console = Console()
        table = Table(title=title, show_lines=True)
        for header, _, _ in cols:
            table.add_column(header, justify="right" if header not in ("场景", "模型") else "left",
                             style="cyan" if header in ("场景", "模型") else None)
        for r in rows:
            cells = []
            for _, getter, fmt in cols:
                cells.append(fmt(getter(r)))
            # 成功率 < 99% 整行标红告警
            if r["success_rate"] < 99:
                cells = [f"[red]{c}[/red]" for c in cells]
            table.add_row(*cells)
        console.print(table)
    else:
        print(f"\n{'='*120}")
        print(f" {title}")
        print(f"{'='*120}")
        headers = [h for h, _, _ in cols]
        # 计算每列最大宽度：对齐表头和所有行
        all_cells = [headers[:]]  # 第一行是表头
        for r in rows:
            all_cells.append([fmt(getter(r)) for _, getter, fmt in cols])
        col_widths = []
        for ci in range(len(headers)):
            max_w = max(len(str(row[ci])) for row in all_cells)
            col_widths.append(max(max_w, len(headers[ci])))
        # 打印表头
        print(" | ".join(f"{headers[i]:>{col_widths[i]}}" for i in range(len(headers))))
        print("-+-".join("-" * w for w in col_widths))
        for row_cells in all_cells[1:]:
            print(" | ".join(f"{str(row_cells[i]):>{col_widths[i]}}" for i in range(len(row_cells))))


def save_comparison(rows: list, model: str, output_dir: str, *, show_model: bool = False):
    """保存对比报告为 Markdown + CSV"""
    now = datetime.now()
    date_dir = Path(output_dir) / now.strftime("%Y-%m-%d")
    date_dir.mkdir(parents=True, exist_ok=True)
    safe_model = model.replace("/", "_").replace(":", "_").replace(" — ", "_")
    stem = f"comparison_{safe_model}_{now.strftime('%H%M%S')}"

    cols = _cols(show_model)
    headers = [h for h, _, _ in cols]

    # ── Markdown ──────────────────────────────────────────
    md_lines = [
        f"# 多场景压测对比 — {model}",
        "",
        f"> 生成时间: {now.strftime('%Y-%m-%d %H:%M:%S')}",
        "",
        "| " + " | ".join(headers) + " |",
        "|" + "|".join(["---"] * len(headers)) + "|",
    ]
    for r in rows:
        cells = [fmt(getter(r)) for _, getter, fmt in cols]
        md_lines.append("| " + " | ".join(cells) + " |")
    md_path = date_dir / f"{stem}.md"
    md_path.write_text("\n".join(md_lines) + "\n", encoding="utf-8")

    # ── CSV ───────────────────────────────────────────────
    csv_path = date_dir / f"{stem}.csv"
    with csv_path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for r in rows:
            # CSV 写原始数值，方便 Excel 计算/画图
            writer.writerow([getter(r) for _, getter, _ in cols])

    print(f"\n💾 对比报告已保存:")
    print(f"   📄 Markdown: {md_path}")
    print(f"   📊 CSV:      {csv_path}")


# ═══════════════════════════════════════════════════
# Excel 导出（面向非技术人员的格式化报表）
# ═══════════════════════════════════════════════════

# 指标的中文说明，用于 Excel 里的"指标说明" sheet
_METRIC_HELP = {
    "模型": "被压测的模型（跨模型对比时用于区分不同模型的同场景结果）",
    "场景": "测试场景的名称",
    "并发": "同时发送请求的数量，数值越高对服务器的压力越大",
    "请求数": "总共发送了多少次请求",
    "入Token": "每次请求平均消耗的输入 Token 数（输入越长数值越大）",
    "出Token": "每次成功请求平均生成的输出 Token 数",
    "总Token": "该场景总共消耗的 Token 数（≈ 花了多少额度）",
    "成功率": "请求成功的比例，100% 表示没有错误发生",
    "QPS": "每秒处理的请求数（Queries Per Second），越高越好",
    "TPS": "每秒生成的 Token 数（Tokens Per Second），反映吞吐能力",
    "延迟Avg": "所有成功请求的平均响应时间，单位秒",
    "延迟Std": "响应时间的标准差，越小表示表现越稳定",
    "延迟P50": "50% 的请求在这个时间内完成（中位数），反映典型用户体验",
    "延迟P95": "95% 的请求在这个时间内完成，反映大多数用户的最差体验",
    "延迟P99": "99% 的请求在这个时间内完成，反映极端情况",
    "TTFT Avg": "首 Token 平均时间（从发送请求到收到第一个字），反映模型思考速度",
    "TTFT P95": "95% 请求的首 Token 时间",
    "TPOT Avg": "每个输出 Token 的平均生成时间，反映模型输出速度",
    "TPOT P95": "95% 请求的每 Token 生成时间",
    "长尾比": "P95延迟÷P50延迟，越接近1.0表示延迟越均匀，>2.0表示存在明显抖动",
}


def save_excel(all_rows: list, groups: dict[str, list], model: str, output_dir: str,
               *, show_model: bool = False):
    """生成格式化的 Excel 报表，适合非技术人员阅读。

    包含：
    - 总览 sheet + 各维度独立 sheet
    - 格式化表头、自动列宽、冻结首行
    - 成功率条件着色、延迟对比柱状图
    - 指标说明 sheet
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
        from openpyxl.utils import get_column_letter
        from openpyxl.chart import BarChart, Reference
    except ImportError:
        print("\n⚠️  需要安装 openpyxl 才能生成 Excel: pip install openpyxl")
        return

    now = datetime.now()
    date_dir = Path(output_dir) / now.strftime("%Y-%m-%d")
    date_dir.mkdir(parents=True, exist_ok=True)
    safe_model = model.replace("/", "_").replace(":", "_").replace(" — ", "_")
    filepath = date_dir / f"report_{safe_model}_{now.strftime('%H%M%S')}.xlsx"

    wb = openpyxl.Workbook()
    # 删除默认空 sheet
    wb.remove(wb.active)

    # ── 样式定义 ──────────────────────────────────
    header_font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    # ── 表头 ──────────────────────────────────────
    cols = _cols(show_model)
    headers = [h for h, _, _ in cols]

    def _write_data_sheet(ws, rows, title):
        """在 worksheet 中写入格式化数据"""
        ws.title = title

        # 写表头
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # 写数据行
        for ri, r in enumerate(rows, 2):
            for ci, (_, getter, fmt) in enumerate(cols, 1):
                value = getter(r)
                cell = ws.cell(row=ri, column=ci, value=fmt(value) if callable(fmt) else value)
                cell.alignment = cell_align
                cell.border = thin_border

        # 列宽自适应
        for ci in range(1, len(headers) + 1):
            max_width = len(headers[ci - 1]) * 2  # 中文字符较宽
            for ri in range(2, len(rows) + 2):
                val = str(ws.cell(row=ri, column=ci).value or "")
                max_width = max(max_width, len(val) * 1.2)
            ws.column_dimensions[get_column_letter(ci)].width = min(max_width + 4, 22)

        # 条件着色：成功率列
        success_col = headers.index("成功率") + 1
        for ri in range(2, len(rows) + 2):
            cell = ws.cell(row=ri, column=success_col)
            try:
                val = float(str(cell.value).replace("%", ""))
                if val < 99:
                    cell.fill = red_fill
                elif val < 100:
                    cell.fill = yellow_fill
                else:
                    cell.fill = green_fill
            except (ValueError, AttributeError):
                pass

        # 冻结首行
        ws.freeze_panes = "A2"

    def _add_chart(ws, rows, title, data_start_row, data_end_row):
        """为 sheet 添加延迟 P50/P95/Avg 对比柱状图"""
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = title
        chart.y_axis.title = "秒"
        chart.x_axis.title = "场景"
        chart.width = min(30, max(15, len(rows) * 3))
        chart.height = 14

        # 场景名在 A 列，延迟 P50/P95/Avg 列
        name_col = headers.index("场景") + 1
        p50_col = headers.index("延迟P50") + 1
        p95_col = headers.index("延迟P95") + 1
        avg_col = headers.index("延迟Avg") + 1

        cats = Reference(ws, min_col=name_col, min_row=2, max_row=data_end_row)
        for label, col in [("P50", p50_col), ("P95", p95_col), ("平均", avg_col)]:
            data_ref = Reference(ws, min_col=col, min_row=1, max_row=data_end_row)
            chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats)

        # 颜色
        if len(chart.series) >= 1:
            chart.series[0].graphicalProperties.solidFill = "5B9BD5"  # P50 蓝
        if len(chart.series) >= 2:
            chart.series[1].graphicalProperties.solidFill = "ED7D31"  # P95 橙
        if len(chart.series) >= 3:
            chart.series[2].graphicalProperties.solidFill = "A5A5A5"  # Avg 灰

        chart_row = data_end_row + 3
        ws.add_chart(chart, f"A{chart_row}")

    # ── 总览 Sheet ────────────────────────────────
    ws_all = wb.create_sheet()
    _write_data_sheet(ws_all, all_rows, "总览")
    _add_chart(ws_all, all_rows, f"延迟对比 — {model}", 2, len(all_rows) + 1)

    # ── 各维度 Sheet ──────────────────────────────
    group_labels = {"输入": "输入长度梯度", "输出": "输出长度梯度", "业务": "业务场景验证"}
    # 先把有名字的排前面
    for prefix in ["输入", "输出", "业务"]:
        if prefix in groups and len(groups[prefix]) >= 2:
            ws = wb.create_sheet()
            label = group_labels.get(prefix, prefix)
            _write_data_sheet(ws, groups[prefix], label)
            _add_chart(ws, groups[prefix], f"延迟对比 — {label}", 2, len(groups[prefix]) + 1)
    # 其余未归类的组
    for prefix, group_rows in sorted(groups.items()):
        if prefix in ("输入", "输出", "业务"):
            continue
        if len(group_rows) >= 2:
            ws = wb.create_sheet()
            _write_data_sheet(ws, group_rows, prefix)
            _add_chart(ws, group_rows, f"延迟对比 — {prefix}", 2, len(group_rows) + 1)

    # ── 指标说明 Sheet ────────────────────────────
    ws_help = wb.create_sheet("指标说明")
    ws_help.column_dimensions["A"].width = 16
    ws_help.column_dimensions["B"].width = 60
    help_header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for ci, text in enumerate(["指标", "说明"], 1):
        cell = ws_help.cell(row=1, column=ci, value=text)
        cell.font = header_font
        cell.fill = help_header_fill
        cell.alignment = header_align
        cell.border = thin_border
    ws_help.freeze_panes = "A2"

    for ri, (metric, desc) in enumerate(_METRIC_HELP.items(), 2):
        cell_a = ws_help.cell(row=ri, column=1, value=metric)
        cell_a.font = Font(name="微软雅黑", bold=True, size=10)
        cell_a.alignment = cell_align
        cell_a.border = thin_border
        cell_b = ws_help.cell(row=ri, column=2, value=desc)
        cell_b.font = Font(name="微软雅黑", size=10)
        cell_b.alignment = Alignment(vertical="center")
        cell_b.border = thin_border

    # ── 保存 ──────────────────────────────────────
    wb.save(filepath)
    print(f"   📗 Excel:     {filepath}")